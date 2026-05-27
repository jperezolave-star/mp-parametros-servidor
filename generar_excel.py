import json, sys, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Datos de ejemplo (se reemplaza con JSON real de la app) ──────────────────
DATOS_EJEMPLO = {
    "edificio": "Torre B - Parque Titanium",
    "mes": "Mayo",
    "anio": "2026",
    "tecnico": "Jose Perez Olave",
    "tableros": [
        {"n":1,"nombre":"3","servicio":"EMERG","piso":3,"v_rs":380,"v_rt":382,"v_st":385,"v_rn":219,"v_sn":220,"v_tn":221,"v_rtie":219,"v_stie":220,"v_ttie":221,"v_ntie":0.2,"i_r":0,"i_s":0,"i_t":0,"i_n":4.4,"i_tie":0.7,"t_r":23,"t_s":23,"t_t":23,"t_n":23,"t_tie":23,"luz_piloto":1,"chapas":0,"obs":"","fecha_hora":"13-05-2026 09:15"},
        {"n":2,"nombre":"4","servicio":"EMERG","piso":4,"v_rs":379,"v_rt":380,"v_st":384,"v_rn":218,"v_sn":219,"v_tn":221,"v_rtie":218,"v_stie":219,"v_ttie":221,"v_ntie":0.3,"i_r":0.1,"i_s":0,"i_t":0.3,"i_n":1.9,"i_tie":0.8,"t_r":25,"t_s":25,"t_t":25,"t_n":25,"t_tie":24,"luz_piloto":0,"chapas":0,"obs":"","fecha_hora":"13-05-2026 09:22"},
        {"n":3,"nombre":"5","servicio":"EMERG","piso":5,"v_rs":380,"v_rt":382,"v_st":384,"v_rn":220,"v_sn":223,"v_tn":222,"v_rtie":220,"v_stie":223,"v_ttie":222,"v_ntie":0.2,"i_r":19.3,"i_s":20.7,"i_t":2.7,"i_n":6.2,"i_tie":0.4,"t_r":28,"t_s":28,"t_t":28,"t_n":28,"t_tie":30,"luz_piloto":0,"chapas":0,"obs":"","fecha_hora":"13-05-2026 09:31"},
        {"n":4,"nombre":"6","servicio":"EMERG","piso":6,"v_rs":381,"v_rt":383,"v_st":385,"v_rn":219,"v_sn":221,"v_tn":222,"v_rtie":219,"v_stie":221,"v_ttie":222,"v_ntie":0.3,"i_r":24.9,"i_s":38.3,"i_t":29.1,"i_n":9.1,"i_tie":0.2,"t_r":28,"t_s":28,"t_t":28,"t_n":26,"t_tie":26,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":5,"nombre":"7","servicio":"EMERG","piso":7,"v_rs":380,"v_rt":381,"v_st":385,"v_rn":219,"v_sn":221,"v_tn":222,"v_rtie":219,"v_stie":221,"v_ttie":222,"v_ntie":0.2,"i_r":23.1,"i_s":35.9,"i_t":20.1,"i_n":2.6,"i_tie":0.8,"t_r":31,"t_s":31,"t_t":31,"t_n":31,"t_tie":31,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":6,"nombre":"8","servicio":"EMERG","piso":8,"v_rs":381,"v_rt":382,"v_st":384,"v_rn":218,"v_sn":220,"v_tn":222,"v_rtie":218,"v_stie":220,"v_ttie":222,"v_ntie":0.7,"i_r":25.6,"i_s":36.5,"i_t":27.7,"i_n":12.9,"i_tie":1.2,"t_r":28,"t_s":28,"t_t":28,"t_n":28,"t_tie":28,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":7,"nombre":"9","servicio":"EMERG","piso":9,"v_rs":380,"v_rt":381,"v_st":385,"v_rn":218,"v_sn":220,"v_tn":222,"v_rtie":218,"v_stie":220,"v_ttie":222,"v_ntie":0.2,"i_r":5.4,"i_s":10.7,"i_t":7.8,"i_n":3.1,"i_tie":2.2,"t_r":27,"t_s":27,"t_t":27,"t_n":27,"t_tie":26,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":8,"nombre":"10","servicio":"EMERG","piso":10,"v_rs":380,"v_rt":382,"v_st":385,"v_rn":219,"v_sn":221,"v_tn":223,"v_rtie":219,"v_stie":221,"v_ttie":223,"v_ntie":0.2,"i_r":21.2,"i_s":23.5,"i_t":26.9,"i_n":1.7,"i_tie":2.2,"t_r":26,"t_s":26,"t_t":26,"t_n":26,"t_tie":26,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":9,"nombre":"11","servicio":"EMERG","piso":11,"v_rs":380,"v_rt":382,"v_st":384,"v_rn":218,"v_sn":221,"v_tn":223,"v_rtie":218,"v_stie":221,"v_ttie":223,"v_ntie":0.2,"i_r":28.8,"i_s":40.8,"i_t":26.9,"i_n":1.9,"i_tie":0.7,"t_r":28,"t_s":28,"t_t":28,"t_n":28,"t_tie":29,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":10,"nombre":"12","servicio":"EMERG","piso":12,"v_rs":380,"v_rt":382,"v_st":384,"v_rn":218,"v_sn":221,"v_tn":223,"v_rtie":218,"v_stie":221,"v_ttie":223,"v_ntie":0.2,"i_r":16.2,"i_s":16.5,"i_t":17.6,"i_n":1.6,"i_tie":0.9,"t_r":26,"t_s":26,"t_t":26,"t_n":25,"t_tie":25,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":11,"nombre":"13","servicio":"EMERG","piso":13,"v_rs":380,"v_rt":382,"v_st":385,"v_rn":220,"v_sn":221,"v_tn":223,"v_rtie":220,"v_stie":221,"v_ttie":223,"v_ntie":0.2,"i_r":6.4,"i_s":6.18,"i_t":6.22,"i_n":2.8,"i_tie":0.3,"t_r":26,"t_s":26,"t_t":26,"t_n":26,"t_tie":26,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":12,"nombre":"14","servicio":"EMERG","piso":14,"v_rs":380,"v_rt":381,"v_st":385,"v_rn":221,"v_sn":220,"v_tn":222,"v_rtie":221,"v_stie":220,"v_ttie":222,"v_ntie":0.2,"i_r":5.6,"i_s":22.1,"i_t":1.8,"i_n":20.1,"i_tie":0,"t_r":26,"t_s":26,"t_t":26,"t_n":26,"t_tie":26,"luz_piloto":0,"chapas":2,"obs":""},
        {"n":13,"nombre":"15","servicio":"EMERG","piso":15,"v_rs":382,"v_rt":384,"v_st":387,"v_rn":220,"v_sn":221,"v_tn":223,"v_rtie":220,"v_stie":221,"v_ttie":223,"v_ntie":0.2,"i_r":14.4,"i_s":23.5,"i_t":12.7,"i_n":4.4,"i_tie":1.6,"t_r":28,"t_s":28,"t_t":28,"t_n":28,"t_tie":28,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":14,"nombre":"16","servicio":"EMERG","piso":16,"v_rs":381,"v_rt":383,"v_st":380,"v_rn":219,"v_sn":220,"v_tn":223,"v_rtie":219,"v_stie":220,"v_ttie":223,"v_ntie":0.2,"i_r":34.4,"i_s":10.4,"i_t":14.4,"i_n":3.4,"i_tie":0.2,"t_r":29,"t_s":29,"t_t":29,"t_n":29,"t_tie":28,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":15,"nombre":"17","servicio":"EMERG","piso":17,"v_rs":382,"v_rt":383,"v_st":386,"v_rn":219,"v_sn":221,"v_tn":224,"v_rtie":219,"v_stie":221,"v_ttie":224,"v_ntie":0.2,"i_r":1.5,"i_s":0.8,"i_t":0,"i_n":1.6,"i_tie":0,"t_r":28,"t_s":28,"t_t":28,"t_n":27,"t_tie":27,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":16,"nombre":"18","servicio":"EMERG","piso":18,"v_rs":380,"v_rt":384,"v_st":385,"v_rn":220,"v_sn":221,"v_tn":224,"v_rtie":220,"v_stie":221,"v_ttie":224,"v_ntie":0.2,"i_r":46.4,"i_s":48.7,"i_t":44.5,"i_n":19.6,"i_tie":0,"t_r":29,"t_s":29,"t_t":29,"t_n":29,"t_tie":29,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":17,"nombre":"19","servicio":"EMERG","piso":19,"v_rs":383,"v_rt":384,"v_st":387,"v_rn":220,"v_sn":221,"v_tn":224,"v_rtie":220,"v_stie":221,"v_ttie":224,"v_ntie":0.2,"i_r":22.3,"i_s":8.9,"i_t":14.6,"i_n":2,"i_tie":1.2,"t_r":28,"t_s":28,"t_t":28,"t_n":26,"t_tie":26,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":18,"nombre":"20","servicio":"EMERG","piso":20,"v_rs":382,"v_rt":384,"v_st":387,"v_rn":220,"v_sn":222,"v_tn":224,"v_rtie":220,"v_stie":222,"v_ttie":224,"v_ntie":0.2,"i_r":11.4,"i_s":12.3,"i_t":10.4,"i_n":0.3,"i_tie":1,"t_r":28,"t_s":28,"t_t":28,"t_n":28,"t_tie":28,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":19,"nombre":"21","servicio":"EMERG","piso":21,"v_rs":381,"v_rt":384,"v_st":386,"v_rn":221,"v_sn":222,"v_tn":224,"v_rtie":221,"v_stie":222,"v_ttie":224,"v_ntie":0.2,"i_r":12.8,"i_s":19.3,"i_t":10.2,"i_n":1.7,"i_tie":1.4,"t_r":28,"t_s":28,"t_t":28,"t_n":28,"t_tie":28,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":20,"nombre":"22","servicio":"EMERG","piso":22,"v_rs":384,"v_rt":386,"v_st":385,"v_rn":220,"v_sn":221,"v_tn":225,"v_rtie":220,"v_stie":221,"v_ttie":225,"v_ntie":0.2,"i_r":14.5,"i_s":11.2,"i_t":14.6,"i_n":1.3,"i_tie":1.2,"t_r":27,"t_s":27,"t_t":27,"t_n":26,"t_tie":26,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":21,"nombre":"23","servicio":"EMERG","piso":23,"v_rs":385,"v_rt":386,"v_st":389,"v_rn":221,"v_sn":222,"v_tn":225,"v_rtie":221,"v_stie":222,"v_ttie":225,"v_ntie":0.2,"i_r":7.5,"i_s":10.4,"i_t":19.6,"i_n":2.9,"i_tie":1.6,"t_r":27,"t_s":27,"t_t":27,"t_n":26,"t_tie":26,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":22,"nombre":"24","servicio":"EMERG","piso":24,"v_rs":385,"v_rt":386,"v_st":389,"v_rn":221,"v_sn":223,"v_tn":225,"v_rtie":221,"v_stie":223,"v_ttie":225,"v_ntie":0.5,"i_r":9.2,"i_s":10.5,"i_t":12,"i_n":10.2,"i_tie":0.6,"t_r":27,"t_s":27,"t_t":27,"t_n":27,"t_tie":27,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":23,"nombre":"TDF y C Torniquetes  closet electrico -2","servicio":"NORM","piso":-2,"v_rs":None,"v_rt":None,"v_st":None,"v_rn":224,"v_sn":None,"v_tn":None,"v_rtie":224,"v_stie":0,"v_ttie":0,"v_ntie":0.3,"i_r":0.9,"i_s":None,"i_t":0.6,"i_n":0.2,"i_tie":25,"t_r":None,"t_s":None,"t_t":25,"t_n":25,"t_tie":0,"luz_piloto":0,"chapas":2,"obs":""},
        {"n":24,"nombre":"TDF - CL- E - 24P","servicio":"EMERG","piso":24,"v_rs":387,"v_rt":388,"v_st":390,"v_rn":222,"v_sn":223,"v_tn":224,"v_rtie":222,"v_stie":223,"v_ttie":224,"v_ntie":0.1,"i_r":5.3,"i_s":5.1,"i_t":4.2,"i_n":1.4,"i_tie":0.6,"t_r":22,"t_s":22,"t_t":22,"t_n":22,"t_tie":22,"luz_piloto":1,"chapas":0,"obs":""},
        {"n":25,"nombre":"TAG TDF - CL 24P","servicio":"NORM","piso":24,"v_rs":396,"v_rt":400,"v_st":400,"v_rn":228,"v_sn":229,"v_tn":230,"v_rtie":228,"v_stie":229,"v_ttie":230,"v_ntie":0.2,"i_r":4.6,"i_s":4.7,"i_t":4.6,"i_n":0.16,"i_tie":0,"t_r":25,"t_s":25,"t_t":25,"t_n":25,"t_tie":25,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":26,"nombre":"TDF Y C CLIMA TORRE B SUBTERRANEO -5","servicio":"NORM","piso":-5,"v_rs":389,"v_rt":391,"v_st":395,"v_rn":228,"v_sn":226,"v_tn":228,"v_rtie":228,"v_stie":226,"v_ttie":228,"v_ntie":0.7,"i_r":45.9,"i_s":52.8,"i_t":55.1,"i_n":0.1,"i_tie":2.4,"t_r":27,"t_s":27,"t_t":27,"t_n":27,"t_tie":27,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":27,"nombre":"TDF - CLIMA 2° PISO NORTE","servicio":"NORM","piso":2,"v_rs":390,"v_rt":392,"v_st":394,"v_rn":225,"v_sn":226,"v_tn":227,"v_rtie":225,"v_stie":226,"v_ttie":227,"v_ntie":0.1,"i_r":13.3,"i_s":14.8,"i_t":16.7,"i_n":0.7,"i_tie":0.1,"t_r":24,"t_s":24,"t_t":24,"t_n":24,"t_tie":24,"luz_piloto":1,"chapas":3,"obs":""},
        {"n":28,"nombre":"TDF - CLIMA 2° PISO SUR","servicio":"NORM","piso":2,"v_rs":390,"v_rt":392,"v_st":394,"v_rn":226,"v_sn":227,"v_tn":227,"v_rtie":226,"v_stie":227,"v_ttie":227,"v_ntie":0.4,"i_r":6.5,"i_s":4.6,"i_t":4.7,"i_n":1.7,"i_tie":0.2,"t_r":22,"t_s":22,"t_t":22,"t_n":22,"t_tie":22,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":29,"nombre":"TDF - CL- N - 24P","servicio":"NORM","piso":24,"v_rs":395,"v_rt":396,"v_st":399,"v_rn":217,"v_sn":228,"v_tn":229,"v_rtie":217,"v_stie":228,"v_ttie":229,"v_ntie":0.1,"i_r":0,"i_s":4.5,"i_t":4.8,"i_n":0.1,"i_tie":1.5,"t_r":22,"t_s":22,"t_t":22,"t_n":21,"t_tie":21,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":30,"nombre":"TDF- VPC - G1 EMERGENCIA P2 SUR PISO 2","servicio":"EMERG","piso":2,"v_rs":380,"v_rt":381,"v_st":385,"v_rn":219,"v_sn":221,"v_tn":222,"v_rtie":219,"v_stie":221,"v_ttie":222,"v_ntie":0.1,"i_r":0,"i_s":0,"i_t":0,"i_n":0,"i_tie":0.1,"t_r":21,"t_s":21,"t_t":21,"t_n":21,"t_tie":21,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":31,"nombre":"TDF- VPC - G2 EMERGENCIA P2 NORTE PISO 2","servicio":"EMERG","piso":2,"v_rs":381,"v_rt":382,"v_st":385,"v_rn":219,"v_sn":221,"v_tn":222,"v_rtie":219,"v_stie":221,"v_ttie":222,"v_ntie":0.1,"i_r":0,"i_s":0,"i_t":0,"i_n":0,"i_tie":0,"t_r":21,"t_s":21,"t_t":21,"t_n":21,"t_tie":21,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":32,"nombre":"ASC A","servicio":"EMERG","piso":24,"v_rs":386,"v_rt":382,"v_st":387,"v_rn":221,"v_sn":222,"v_tn":224,"v_rtie":221,"v_stie":222,"v_ttie":224,"v_ntie":0.4,"i_r":0.9,"i_s":3.13,"i_t":6.2,"i_n":6.1,"i_tie":0.1,"t_r":23,"t_s":23,"t_t":23,"t_n":23,"t_tie":23,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":33,"nombre":"ASC B","servicio":"EMERG","piso":24,"v_rs":384,"v_rt":385,"v_st":388,"v_rn":221,"v_sn":223,"v_tn":225,"v_rtie":221,"v_stie":223,"v_ttie":225,"v_ntie":0.14,"i_r":3.7,"i_s":5.4,"i_t":4.7,"i_n":0.7,"i_tie":0.7,"t_r":23,"t_s":23,"t_t":23,"t_n":23,"t_tie":23,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":34,"nombre":"ASC C","servicio":"EMERG","piso":24,"v_rs":387,"v_rt":386,"v_st":385,"v_rn":221,"v_sn":222,"v_tn":223,"v_rtie":221,"v_stie":222,"v_ttie":223,"v_ntie":0.1,"i_r":3.8,"i_s":4.6,"i_t":1.3,"i_n":1.2,"i_tie":0,"t_r":24,"t_s":24,"t_t":24,"t_n":24,"t_tie":24,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":35,"nombre":"ASC D","servicio":"EMERG","piso":24,"v_rs":384,"v_rt":386,"v_st":387,"v_rn":222,"v_sn":223,"v_tn":225,"v_rtie":222,"v_stie":223,"v_ttie":225,"v_ntie":0.1,"i_r":1.8,"i_s":4.6,"i_t":7.2,"i_n":0.5,"i_tie":0.1,"t_r":23,"t_s":23,"t_t":23,"t_n":23,"t_tie":23,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":36,"nombre":"ASC E","servicio":"EMERG","piso":16,"v_rs":402,"v_rt":386,"v_st":388,"v_rn":221,"v_sn":222,"v_tn":224,"v_rtie":221,"v_stie":222,"v_ttie":224,"v_ntie":0.3,"i_r":4.6,"i_s":1.5,"i_t":3,"i_n":1.2,"i_tie":0.3,"t_r":25,"t_s":25,"t_t":25,"t_n":25,"t_tie":25,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":37,"nombre":"ASC F","servicio":"EMERG","piso":16,"v_rs":384,"v_rt":388,"v_st":385,"v_rn":223,"v_sn":224,"v_tn":223,"v_rtie":223,"v_stie":224,"v_ttie":223,"v_ntie":0.7,"i_r":1.4,"i_s":1.6,"i_t":1.4,"i_n":0,"i_tie":0.2,"t_r":24,"t_s":24,"t_t":24,"t_n":24,"t_tie":24,"luz_piloto":1,"chapas":0,"obs":""},
        {"n":38,"nombre":"ASC G","servicio":"EMERG","piso":16,"v_rs":383,"v_rt":385,"v_st":387,"v_rn":221,"v_sn":222,"v_tn":223,"v_rtie":221,"v_stie":222,"v_ttie":223,"v_ntie":0.1,"i_r":1.2,"i_s":1.5,"i_t":3.7,"i_n":1.3,"i_tie":0.3,"t_r":24,"t_s":24,"t_t":24,"t_n":24,"t_tie":24,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":39,"nombre":"ASC H","servicio":"EMERG","piso":16,"v_rs":383,"v_rt":385,"v_st":387,"v_rn":220,"v_sn":222,"v_tn":224,"v_rtie":220,"v_stie":222,"v_ttie":224,"v_ntie":0.4,"i_r":0.4,"i_s":1.5,"i_t":1.3,"i_n":0.4,"i_tie":0,"t_r":24,"t_s":24,"t_t":24,"t_n":24,"t_tie":24,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":40,"nombre":"TDF  AUX ASCENSORES SUBTERRÁNEO","servicio":"EMERG","piso":-3,"v_rs":382,"v_rt":383,"v_st":388,"v_rn":220,"v_sn":222,"v_tn":224,"v_rtie":220,"v_stie":222,"v_ttie":224,"v_ntie":0.8,"i_r":20.1,"i_s":13.7,"i_t":13.8,"i_n":0.9,"i_tie":1,"t_r":22,"t_s":22,"t_t":22,"t_n":22,"t_tie":22,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":41,"nombre":"TDAyF LOCALES COMERCIALES 1°SUBT. EMERGENCIA","servicio":"EMERG","piso":-2,"v_rs":380,"v_rt":382,"v_st":385,"v_rn":219,"v_sn":221,"v_tn":222,"v_rtie":219,"v_stie":221,"v_ttie":222,"v_ntie":0.1,"i_r":0.5,"i_s":13.3,"i_t":0.1,"i_n":17.3,"i_tie":6.2,"t_r":23,"t_s":23,"t_t":23,"t_n":23,"t_tie":23,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":42,"nombre":"TDAyF TORRE B PQ2 NORMAL","servicio":"NORM","piso":-2,"v_rs":381,"v_rt":383,"v_st":385,"v_rn":220,"v_sn":222,"v_tn":222,"v_rtie":220,"v_stie":222,"v_ttie":222,"v_ntie":0.3,"i_r":2.4,"i_s":9.2,"i_t":4.5,"i_n":10.6,"i_tie":3.6,"t_r":22,"t_s":22,"t_t":22,"t_n":22,"t_tie":22,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":43,"nombre":"TDAyF TORRE B PQ2 EMERGENCIA","servicio":"EMERG","piso":-2,"v_rs":380,"v_rt":381,"v_st":385,"v_rn":220,"v_sn":221,"v_tn":223,"v_rtie":220,"v_stie":221,"v_ttie":223,"v_ntie":0.2,"i_r":13.2,"i_s":29.1,"i_t":17.8,"i_n":31.3,"i_tie":4.9,"t_r":24,"t_s":24,"t_t":24,"t_n":24,"t_tie":24,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":44,"nombre":"TGAFyCOMP. EMERGENCIA","servicio":"EMERG","piso":-2,"v_rs":382,"v_rt":385,"v_st":384,"v_rn":219,"v_sn":220,"v_tn":222,"v_rtie":219,"v_stie":220,"v_ttie":222,"v_ntie":0.5,"i_r":90,"i_s":105,"i_t":100,"i_n":42.3,"i_tie":7.5,"t_r":26,"t_s":26,"t_t":26,"t_n":26,"t_tie":24,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":45,"nombre":"TGAFyCOMP. NORMAL   trafo 1","servicio":"NORM","piso":-2,"v_rs":382,"v_rt":385,"v_st":384,"v_rn":219,"v_sn":220,"v_tn":222,"v_rtie":219,"v_stie":220,"v_ttie":222,"v_ntie":0.5,"i_r":743,"i_s":652,"i_t":653,"i_n":286,"i_tie":0,"t_r":24,"t_s":24,"t_t":24,"t_n":24,"t_tie":24,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":46,"nombre":"TG.Aux. AyF local COMERCIAL 1° SUBT. NORMAL","servicio":"NORM","piso":-2,"v_rs":392,"v_rt":395,"v_st":393,"v_rn":226,"v_sn":227,"v_tn":228,"v_rtie":226,"v_stie":227,"v_ttie":228,"v_ntie":0.1,"i_r":45.8,"i_s":45.2,"i_t":48.9,"i_n":12.7,"i_tie":1,"t_r":20,"t_s":20,"t_t":20,"t_n":20,"t_tie":20,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":47,"nombre":"TGAFyCOMP. NORMAL TRAFO #2","servicio":"NORM","piso":-2,"v_rs":391,"v_rt":393,"v_st":395,"v_rn":224,"v_sn":227,"v_tn":228,"v_rtie":224,"v_stie":227,"v_ttie":228,"v_ntie":0,"i_r":71.8,"i_s":45.4,"i_t":54.2,"i_n":0,"i_tie":0,"t_r":24,"t_s":24,"t_t":24,"t_n":24,"t_tie":24,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":48,"nombre":"BANCO DE CONDENSADOR N° 1","servicio":"NORM","piso":-2,"v_rs":382,"v_rt":387,"v_st":384,"v_rn":None,"v_sn":None,"v_tn":None,"v_rtie":None,"v_stie":None,"v_ttie":None,"v_ntie":None,"i_r":0.82,"i_s":None,"i_t":None,"i_n":None,"i_tie":None,"t_r":None,"t_s":None,"t_t":None,"t_n":None,"t_tie":None,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":49,"nombre":"BANCO DE CONDENSADOR N° 2","servicio":"NORM","piso":-2,"v_rs":392,"v_rt":397,"v_st":393,"v_rn":None,"v_sn":None,"v_tn":None,"v_rtie":None,"v_stie":None,"v_ttie":None,"v_ntie":None,"i_r":162,"i_s":None,"i_t":None,"i_n":None,"i_tie":None,"t_r":None,"t_s":None,"t_t":None,"t_n":None,"t_tie":None,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":50,"nombre":"TABLERO SALA REPLICA","servicio":"NORM","piso":-4,"v_rs":386,"v_rt":387,"v_st":390,"v_rn":222,"v_sn":224,"v_tn":225,"v_rtie":222,"v_stie":224,"v_ttie":225,"v_ntie":0.2,"i_r":1.3,"i_s":1.5,"i_t":1.3,"i_n":1.4,"i_tie":0.2,"t_r":22,"t_s":22,"t_t":22,"t_n":22,"t_tie":22,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":51,"nombre":"Td comp sala seguridad  sscc","servicio":"EMERG","piso":-2,"v_rs":382,"v_rt":383,"v_st":386,"v_rn":221,"v_sn":223,"v_tn":225,"v_rtie":221,"v_stie":223,"v_ttie":225,"v_ntie":0.4,"i_r":5.4,"i_s":5.5,"i_t":5.3,"i_n":2.8,"i_tie":0.4,"t_r":22,"t_s":22,"t_t":22,"t_n":22,"t_tie":22,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":52,"nombre":"DUCTO DE BARRA TB","servicio":"EMERG","piso":-2,"v_rs":381,"v_rt":382,"v_st":385,"v_rn":219,"v_sn":221,"v_tn":223,"v_rtie":219,"v_stie":221,"v_ttie":223,"v_ntie":0,"i_r":118,"i_s":111.7,"i_t":108.8,"i_n":0.1,"i_tie":0.1,"t_r":25,"t_s":25,"t_t":25,"t_n":25,"t_tie":25,"luz_piloto":0,"chapas":0,"obs":""},
        {"n":53,"nombre":"TABLERO AGUA POTABLE TB","servicio":"EMERG","piso":-5,"v_rs":381,"v_rt":383,"v_st":387,"v_rn":220,"v_sn":222,"v_tn":223,"v_rtie":None,"v_stie":220,"v_ttie":222,"v_ntie":0.3,"i_r":5.11,"i_s":6.9,"i_t":4.8,"i_n":0.2,"i_tie":0.3,"t_r":29,"t_s":29,"t_t":29,"t_n":29,"t_tie":28,"luz_piloto":0,"chapas":0,"obs":""}
    ]
}

# Mapa tablero N° → fila Excel y categorías separadoras
# Estructura: lista de items: ('cat', nombre) o ('tb', n)
ESTRUCTURA = [
    ('tb', range(1, 24)),       # Pisos 3-24 + TDF Torniquetes
    ('cat', 'CLIMA'),
    ('tb', range(24, 30)),      # Clima
    ('cat', 'PREZURIZACION '),
    ('tb', range(30, 32)),      # Presurización
    ('cat', 'ASCENSORES '),
    ('tb', range(32, 41)),      # Ascensores
    ('cat', 'SALA ELECTRICA'),
    ('tb', range(41, 53)),      # Sala eléctrica
    ('cat', 'AGUA POTABLE '),
    ('tb', range(53, 54)),      # Agua potable
]

def val(v):
    return v if v is not None else None

def build_excel(datos, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = datos['mes'].lower()

    # ── Estilos ─────────────────────────────────────────────────────────────
    fill_header_title = PatternFill("solid", fgColor="4472C4")  # azul oscuro similar al original
    fill_green_cat    = PatternFill("solid", fgColor="CCFFCC")  # verde categoría
    fill_tensiones    = PatternFill("solid", fgColor="DDEBF7")  # azul claro
    fill_corrientes   = PatternFill("solid", fgColor="E2EFDA")  # verde claro
    fill_temperaturas = PatternFill("solid", fgColor="FFF2CC")  # amarillo claro

    font_title  = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    font_header = Font(name="Arial", bold=True, size=9)
    font_cat    = Font(name="Arial", bold=True, size=10)
    font_data   = Font(name="Arial", size=9)

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Anchos de columnas ──────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 10.5
    ws.column_dimensions['B'].width = 55.4
    ws.column_dimensions['C'].width = 12.3
    ws.column_dimensions['D'].width = 10.5
    for col in 'EFGHIJKLMNOPQRSTUVWXYZ':
        ws.column_dimensions[col].width = 13.0
    ws.column_dimensions['AA'].width = 18.0

    # ── Fila 1-3: vacías con altura mínima ──────────────────────────────────
    for r in [1,2,3]:
        ws.row_dimensions[r].height = 14.4

    # ── Fila 4: Título edificio (merge A4:Z7) ───────────────────────────────
    ws.merge_cells('A4:Z7')
    c = ws['A4']
    c.value = datos['edificio'].upper()
    c.font  = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    c.fill  = fill_header_title
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 35

    # ── Filas 8-10: Cabeceras ────────────────────────────────────────────────
    ws.row_dimensions[8].height  = 14.4
    ws.row_dimensions[9].height  = 30
    ws.row_dimensions[10].height = 15

    # Merge A8:A10, B8:B10, C8:C10, D8:D10
    for col in ['A','B','C','D']:
        ws.merge_cells(f'{col}8:{col}10')

    # Grupos de cabeceras
    ws.merge_cells('E8:N8')
    ws.merge_cells('O8:S8')
    ws.merge_cells('T8:X8')
    ws.merge_cells('Y8:AA8')

    headers_group = {
        'A8': ('Nº',           fill_tensiones,    align_center),
        'B8': ('TORRE B',      fill_tensiones,    align_center),
        'C8': ('SERVICIO',     fill_tensiones,    align_center),
        'D8': ('PISO',         fill_tensiones,    align_center),
        'E8': ('TENSIONES',    fill_tensiones,    align_center),
        'O8': ('CORRIENTES',   fill_corrientes,   align_center),
        'T8': ('TEMPERATURAS', fill_temperaturas, align_center),
        'Y8': ('OBSERVACIONES',fill_tensiones,    align_center),
    }
    for coord, (val_h, fill_h, aln_h) in headers_group.items():
        c = ws[coord]
        c.value = val_h
        c.font  = font_header
        c.fill  = fill_h
        c.alignment = aln_h
        c.border = border

    # Sub-cabeceras fila 9 (cada una merge col9:col10)
    sub_headers = [
        ('E','FASE R/S',fill_tensiones),('F','FASE R/T',fill_tensiones),
        ('G','FASE S/T',fill_tensiones),('H','FASE R/N',fill_tensiones),
        ('I','FASE S/N',fill_tensiones),('J','FASE T/N',fill_tensiones),
        ('K','FASE R/TIE',fill_tensiones),('L','FASE S/TIE',fill_tensiones),
        ('M','FASE T/TIE',fill_tensiones),('N','NEU/TIE',fill_tensiones),
        ('O','CTE R',fill_corrientes),('P','CTE S',fill_corrientes),
        ('Q','CTE T',fill_corrientes),('R','CTE N',fill_corrientes),
        ('S','CTE TIE',fill_corrientes),
        ('T','Tº R',fill_temperaturas),('U','Tº S',fill_temperaturas),
        ('V','Tº T',fill_temperaturas),('W','Tº N',fill_temperaturas),
        ('X','Tº TIE',fill_temperaturas),
        ('Y','luz piloto',fill_tensiones),('Z','chapas',fill_tensiones),
        ('AA','FECHA / HORA',fill_tensiones),
    ]
    for col_l, label, fill_h in sub_headers:
        ws.merge_cells(f'{col_l}9:{col_l}10')
        c = ws[f'{col_l}9']
        c.value = label
        c.font  = font_header
        c.fill  = fill_h
        c.alignment = align_center
        c.border = border

    # ── Datos ────────────────────────────────────────────────────────────────
    by_n = {t['n']: t for t in datos['tableros']}
    row = 11

    for item in ESTRUCTURA:
        kind, val_item = item[0], item[1]

        if kind == 'cat':
            ws.merge_cells(f'A{row}:AA{row}')
            c = ws[f'A{row}']
            c.value = val_item
            c.font  = font_cat
            c.fill  = fill_green_cat
            c.alignment = align_center
            c.border = border
            ws.row_dimensions[row].height = 20
            row += 1

        else:  # 'tb'
            for n in val_item:
                tb = by_n.get(n)
                ws.row_dimensions[row].height = 43.5
                if not tb:
                    row += 1
                    continue

                def sc(col_l, value, aln=None):
                    c = ws[f'{col_l}{row}']
                    c.value = value
                    c.font  = font_data
                    c.alignment = aln or align_center
                    c.border = border

                sc('A', tb['n'])
                sc('B', tb['nombre'], align_left)
                sc('C', tb['servicio'])
                sc('D', tb['piso'])
                sc('E', val(tb.get('v_rs')))
                sc('F', val(tb.get('v_rt')))
                sc('G', val(tb.get('v_st')))
                sc('H', val(tb.get('v_rn')))
                sc('I', val(tb.get('v_sn')))
                sc('J', val(tb.get('v_tn')))
                sc('K', val(tb.get('v_rtie')))
                sc('L', val(tb.get('v_stie')))
                sc('M', val(tb.get('v_ttie')))
                sc('N', val(tb.get('v_ntie')))
                sc('O', val(tb.get('i_r')))
                sc('P', val(tb.get('i_s')))
                sc('Q', val(tb.get('i_t')))
                sc('R', val(tb.get('i_n')))
                sc('S', val(tb.get('i_tie')))
                sc('T', val(tb.get('t_r')))
                sc('U', val(tb.get('t_s')))
                sc('V', val(tb.get('t_t')))
                sc('W', val(tb.get('t_n')))
                sc('X', val(tb.get('t_tie')))
                sc('Y', val(tb.get('luz_piloto')))
                sc('Z', val(tb.get('chapas')))
                # Fecha/hora de registro
                fecha_hora = tb.get('fecha_hora')
                c_fh = ws[f'AA{row}']
                c_fh.value = fecha_hora if fecha_hora else None
                c_fh.font  = font_data
                c_fh.alignment = align_center
                c_fh.border = border
                row += 1

    # ── Fila final: técnico y fecha ──────────────────────────────────────────
    ws.merge_cells(f'A{row}:AA{row}')
    c = ws[f'A{row}']
    c.value = f"Técnico: {datos['tecnico']}    |    Período: {datos['mes']} {datos['anio']}"
    c.font  = Font(name="Arial", bold=True, size=9)
    c.fill  = fill_green_cat
    c.alignment = align_center
    c.border = border
    ws.row_dimensions[row].height = 18

    # ── Congelar filas de cabecera ───────────────────────────────────────────
    ws.freeze_panes = 'A11'

    wb.save(output_path)
    print(f"Excel guardado en: {output_path}")

# ── Main ─────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    if len(sys.argv) > 1:
        with open(sys.argv[1]) as f:
            datos = json.load(f)
    else:
        datos = DATOS_EJEMPLO

    if len(sys.argv) > 2:
        output = sys.argv[2]
    else:
        mes = datos['mes'].upper()
        anio = datos['anio']
        edificio_slug = datos['edificio'].replace(' ','_').replace('-','').replace('/','')[:20]
        output = f"/mnt/user-data/outputs/PARAMETROS_{edificio_slug}_{mes}_{anio}.xlsx"

    os.makedirs(os.path.dirname(output), exist_ok=True)
    build_excel(datos, output)
